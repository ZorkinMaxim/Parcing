import requests
from lxml import etree
import lxml.html
import openpyxl

def parse(url):
    api = requests.get(url)
    tree = lxml.html.document_fromstring(api.text)
    links = tree.xpath('//*[@id="content"]//p/text()')
    return links

def main():
    url = "https://unicom.md/televizory/?limit=100"
    wb = openpyxl.Workbook()
    link = parse(url)
    wb.create_sheet(title='Page 1', index=0)
    sheet = wb['Page 1']
    for i in link:
        cell = sheet.cell(row=link.index(i) + 1, column=1)
        cell.value = i
    wb.save('Unicom_Televizory.xls')


    # i = 2
    # wb = openpyxl.Workbook()
    # wb.remove(wb['Sheet'])
    # while i <= 12:
    #     words = parse(url.format(number=i))
    #     wb.create_sheet(title=f'Page{i}')
    #     sheet = wb[f'Page{i}']
    #     for word in words:
    #         cell = sheet.cell(row=words.index(word) + 1, column=1)
    #         cell.value = word
    #     i += 1
    # wb.save('All-possible-words.xls')

if __name__ == "__main__":
    main()