import requests
from lxml import etree
import lxml.html
import openpyxl

def parse(url):
    api = requests.get(url)
    tree = lxml.html.document_fromstring(api.text)
    words = tree.xpath('/html/body/div[4]/div[1]/div[1]/div[1]/div[2]/ul/li/a/text()')
    return words

def main():
    url = "http://www.allscrabblewords.com/{number}-letter-words/"
    i = 2
    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    while i <= 12:
        words = parse(url.format(number=i))
        wb.create_sheet(title=f'Page{i}')
        sheet = wb[f'Page{i}']
        for word in words:
            cell = sheet.cell(row=words.index(word) + 1, column=1)
            cell.value = word
        i += 1
    wb.save('All-possible-words.xls')

if __name__ == "__main__":
    main()